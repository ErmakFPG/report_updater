import openpyxl
from openpyxl.styles import PatternFill
from worksheetNumber import WorksheetNumber

MAIN_FILE_NAME = 'ремонт_2021.xlsx'
CHILD_FILE_NAME = 'ГЗ_ТА.xlsx'
ROWS_NUMBER_TO_CHECK = 100


def are_names_same(name1, name2):
    try:
        name1 = "".join(name1.split())
        name2 = "".join(name2.split())
    except AttributeError:
        return False
    return name1 == name2


def update_report(main_worksheet, child_worksheet, column_cost="G"):
    index = 1
    for i in range(1, ROWS_NUMBER_TO_CHECK):
        if child_worksheet["A" + str(i)].value == index or child_worksheet["A" + str(i)].value == str(index) + '.':
            index += 1
            current_child_event = child_worksheet["B" + str(i)].value
            for j in range(1, ROWS_NUMBER_TO_CHECK):
                current_main_event = main_worksheet["B" + str(j)].value
                if are_names_same(current_main_event, current_child_event):
                    main_worksheet["E" + str(j)] = child_worksheet[column_cost + str(i)].value
                    child_worksheet["A" + str(i)].fill =\
                        PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    main_worksheet["A" + str(j)].fill =\
                        PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    break


def main():
    main_workbook = openpyxl.load_workbook(MAIN_FILE_NAME)
    main_worksheet = main_workbook.worksheets[WorksheetNumber.MAIN.value]
    child_workbook = openpyxl.load_workbook(CHILD_FILE_NAME)
    child_worksheet_repair = child_workbook.worksheets[WorksheetNumber.REPAIR.value]
    child_worksheet_repair_bkad = child_workbook.worksheets[WorksheetNumber.REPAIR_BKAD.value]
    child_worksheet_bridges = child_workbook.worksheets[WorksheetNumber.REPAIR_BRIDGES.value]

    update_report(main_worksheet, child_worksheet_repair)
    update_report(main_worksheet, child_worksheet_repair_bkad)
    update_report(main_worksheet, child_worksheet_bridges, column_cost="F")

    main_workbook.save(MAIN_FILE_NAME)
    child_workbook.save(CHILD_FILE_NAME)


if __name__ == '__main__':
    main()
